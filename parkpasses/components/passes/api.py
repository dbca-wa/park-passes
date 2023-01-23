import logging
import pickle
import re
import sys
import uuid
from decimal import Decimal

import openpyxl
import requests
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.http import FileResponse, Http404
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from ledger_api_client.ledger_models import EmailUserRO as EmailUser
from ledger_api_client.utils import create_basket_session, create_checkout_session
from org_model_logs.models import UserAction
from rest_framework import generics, mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import BrowsableAPIRenderer, JSONRenderer
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle
from rest_framework.views import APIView
from rest_framework_datatables.filters import DatatablesFilterBackend
from rest_framework_datatables.pagination import DatatablesPageNumberPagination

from parkpasses.components.cart.models import Cart, CartItem
from parkpasses.components.cart.utils import CartUtils
from parkpasses.components.concessions.models import Concession, ConcessionUsage
from parkpasses.components.discount_codes.models import DiscountCode, DiscountCodeUsage
from parkpasses.components.main.api import (
    CustomDatatablesListMixin,
    CustomDatatablesRenderer,
    UserActionViewSet,
)
from parkpasses.components.main.serializers import UserActionSerializer
from parkpasses.components.orders.models import Order, OrderItem
from parkpasses.components.passes.exceptions import NoValidPassTypeFoundInPost
from parkpasses.components.passes.models import (
    DistrictPassTypeDurationOracleCode,
    Pass,
    PassTemplate,
    PassType,
    PassTypePricingWindow,
    PassTypePricingWindowOption,
    RACDiscountUsage,
)
from parkpasses.components.passes.serializers import (
    ExternalCreateAllParksPassSerializer,
    ExternalCreateAnnualLocalPassSerializer,
    ExternalCreateDayEntryPassSerializer,
    ExternalCreateGoldStarPassSerializer,
    ExternalCreateHolidayPassSerializer,
    ExternalCreatePinjarOffRoadPassSerializer,
    ExternalPassSerializer,
    ExternalUpdatePassSerializer,
    InternalCreatePricingWindowSerializer,
    InternalDistrictPassTypeDurationOracleCodeListUpdateSerializer,
    InternalDistrictPassTypeDurationOracleCodeSerializer,
    InternalOptionSerializer,
    InternalPassCancellationSerializer,
    InternalPassRetrieveSerializer,
    InternalPassSerializer,
    InternalPassTypeSerializer,
    InternalPricingWindowSerializer,
    OptionSerializer,
    PassTemplateSerializer,
    PassTypeSerializer,
    RetailerApiCreatePassSerializer,
    RetailerUpdatePassSerializer,
)
from parkpasses.components.retailers.models import (
    RetailerGroup,
    RetailerGroupAPIKey,
    RetailerGroupUser,
)
from parkpasses.components.vouchers.models import Voucher, VoucherTransaction
from parkpasses.helpers import (
    check_rac_discount_hash,
    get_rac_discount_code,
    get_retailer_group_ids_for_user,
    is_customer,
    is_internal,
    is_retailer,
)
from parkpasses.permissions import (
    HasRetailerGroupAPIKey,
    IsInternal,
    IsInternalAPIView,
    IsInternalDestroyer,
    IsRetailer,
)

# from rest_framework_datatables.filters import DatatablesFilterBackend


logger = logging.getLogger(__name__)


class PassTypesDistinct(APIView):
    permission_classes = [IsAuthenticated]

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def get(self, request, format=None):
        pass_types = [
            {"code": pass_type.id, "description": pass_type.display_name}
            for pass_type in PassType.objects.all().distinct()
        ]
        return Response(pass_types)


class RetailerPassTypesDistinct(APIView):
    permission_classes = [IsAuthenticated]

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def get(self, request, format=None):
        pass_types = [
            {"code": pass_type.id, "description": pass_type.display_name}
            for pass_type in PassType.objects.filter(display_retailer=True)
        ]
        return Response(pass_types)


class PassProcessingStatusesDistinct(APIView):
    permission_classes = [IsAuthenticated]

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def get(self, request, format=None):
        if request.query_params.get("for_filter", ""):
            processing_status_choices = [
                {"id": processing_status_choice[0], "text": processing_status_choice[1]}
                for processing_status_choice in Pass.PROCESSING_STATUS_CHOICES
            ]
        else:
            processing_status_choices = [
                {
                    "code": processing_status_choice[0],
                    "description": processing_status_choice[1],
                }
                for processing_status_choice in Pass.PROCESSING_STATUS_CHOICES
            ]
        return Response(processing_status_choices)


class ExternalPassTypeViewSet(
    mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet
):
    lookup_field = "slug"
    model = PassType
    queryset = PassType.objects.filter(display_externally=True)
    ordering = "display_order"
    serializer_class = PassTypeSerializer

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def retrieve(self, request, slug=None):
        return super().retrieve(request, slug=slug)

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def list(self, request, slug=None):
        return super().list(request, slug=slug)


class RetailerPassTypeViewSet(
    mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet
):
    lookup_field = "slug"
    model = PassType
    queryset = PassType.objects.filter(display_retailer=True)
    ordering = "display_order"
    serializer_class = PassTypeSerializer
    permission_classes = [IsRetailer]

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def retrieve(self, request, slug=None):
        return super().retrieve(request, slug=slug)

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def list(self, request, slug=None):
        logger.debug("RetailerPassTypeViewSet.list")
        return super().list(request, slug=slug)


class InternalPassTypeViewSet(viewsets.ModelViewSet):
    """
    A ViewSet for performing actions on pass types.
    """

    lookup_field = "slug"
    model = PassType
    queryset = PassType.objects.all()
    ordering = "display_order"
    permission_classes = [IsInternal]
    serializer_class = InternalPassTypeSerializer

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def retrieve(self, request, slug=None):
        return super().retrieve(request, slug=slug)

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def list(self, request, slug=None):
        return super().list(request, slug=slug)


class PricingWindowFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):
        total_count = queryset.count()

        pass_type = request.GET.get("pass_type")

        start_date_from = request.GET.get("start_date_from")
        start_date_to = request.GET.get("start_date_to")

        expiry_date_from = request.GET.get("expiry_date_from")
        expiry_date_to = request.GET.get("expiry_date_to")

        if pass_type:
            queryset = queryset.filter(pass_type__id=pass_type)

        if start_date_from:
            queryset = queryset.filter(date_start__gte=start_date_from)

        if start_date_to:
            queryset = queryset.filter(date_start__lte=start_date_to)

        if expiry_date_from:
            queryset = queryset.filter(date_expiry__gte=expiry_date_from)

        if expiry_date_to:
            queryset = queryset.filter(date_expiry__lte=expiry_date_to)

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        queryset = super().filter_queryset(request, queryset, view)
        setattr(view, "_datatables_total_count", total_count)

        return queryset


class InternalPricingWindowViewSet(CustomDatatablesListMixin, viewsets.ModelViewSet):
    search_fields = [
        "name",
    ]
    model = PassTypePricingWindow
    pagination_class = DatatablesPageNumberPagination
    queryset = PassTypePricingWindow.objects.all()
    permission_classes = [IsInternalDestroyer]
    filter_backends = (PricingWindowFilterBackend,)
    renderer_classes = (JSONRenderer, BrowsableAPIRenderer, CustomDatatablesRenderer)

    def get_serializer_class(self):
        if "create" == self.action:
            return InternalCreatePricingWindowSerializer
        return InternalPricingWindowSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.date_start > timezone.now().date():
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response(status=status.HTTP_405_METHOD_NOT_ALLOWED)


class CurrentOptionsForPassType(generics.ListAPIView):
    serializer_class = OptionSerializer

    def get_queryset(self):
        pass_type_id = self.request.query_params.get("pass_type_id")
        options = PassTypePricingWindowOption.get_current_options_by_pass_type_id(
            int(pass_type_id)
        )
        if options:
            return options
        return PassTypePricingWindowOption.objects.none()

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def get(self, *args, **kwargs):
        return super().get(*args, **kwargs)


class DefaultOptionsForPassType(generics.ListAPIView):
    """Updated docstring"""

    permission_classes = [IsInternalAPIView]
    serializer_class = OptionSerializer

    def get_queryset(self):
        pass_type_id = self.request.query_params.get("pass_type_id")
        if pass_type_id.isnumeric():
            return PassTypePricingWindowOption.get_default_options_by_pass_type_id(
                pass_type_id
            )
        return PassTypePricingWindowOption.objects.none()

    @method_decorator(cache_page(settings.CACHE_TIMEOUT_2_HOURS))
    def get(self, *args, **kwargs):
        return super().get(*args, **kwargs)


class PassTypePricingWindowOptionViewSet(viewsets.ModelViewSet):
    """
    A ViewSet for performing actions on pricing windows options.
    """

    model = PassTypePricingWindowOption
    serializer_class = OptionSerializer

    def get_queryset(self):
        return PassTypePricingWindowOption.objects.all()

    def get_serializer_class(self):
        if is_internal(self.request):
            return InternalOptionSerializer
        elif is_retailer(self.request):
            return OptionSerializer
        else:
            return OptionSerializer

    def has_permission(self, request, view):
        if is_internal(request):
            return True
        if is_retailer(self.request):
            if view.action in [
                "list",
                "retrieve",
            ]:
                return True
            return False
        if is_customer(request):
            if view.action in [
                "list",
                "retrieve",
            ]:
                return True
        return False


class PassTemplateViewSet(viewsets.ModelViewSet):
    """
    A ViewSet for performing actions on pass templates.
    """

    model = PassTemplate
    queryset = PassTemplate.objects.all()
    serializer_class = PassTemplateSerializer
    permission_classes = [IsInternal]

    @action(methods=["GET"], detail=True, url_path="retrieve-pass-template")
    def retrieve_pass_template(self, request, *args, **kwargs):
        pass_template = self.get_object()
        if pass_template.template:
            return FileResponse(pass_template.template)
        raise Http404


class SmallResultSetPagination(DatatablesPageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 1000


class ExternalPassViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    viewsets.GenericViewSet,
):
    pagination_class = SmallResultSetPagination
    model = Pass

    def get_queryset(self):
        return (
            Pass.objects.exclude(user__isnull=True)
            .exclude(processing_status=Pass.CANCELLED)
            .exclude(in_cart=True)
            .filter(user=self.request.user.id)
            .order_by("-date_start")
        )

    def get_serializer_class(self):
        if self.action in ["update", "partial_update"]:
            return ExternalUpdatePassSerializer
        if "create" == self.action:
            if "pass_type_name" in self.request.data:
                pass_type_name = self.request.data["pass_type_name"]
                if pass_type_name:
                    if settings.HOLIDAY_PASS == pass_type_name:
                        return ExternalCreateHolidayPassSerializer
                    if settings.ANNUAL_LOCAL_PASS == pass_type_name:
                        return ExternalCreateAnnualLocalPassSerializer
                    if settings.ALL_PARKS_PASS == pass_type_name:
                        return ExternalCreateAllParksPassSerializer
                    if settings.GOLD_STAR_PASS == pass_type_name:
                        return ExternalCreateGoldStarPassSerializer
                    if settings.DAY_ENTRY_PASS == pass_type_name:
                        return ExternalCreateDayEntryPassSerializer
                    if (
                        settings.PINJAR_OFF_ROAD_VEHICLE_AREA_ANNUAL_PASS
                        == pass_type_name
                    ):
                        return ExternalCreatePinjarOffRoadPassSerializer

                    error = "ERROR: No valid pass type name found in POST."
                    logger.error(error)
                    raise NoValidPassTypeFoundInPost(error)

        return ExternalPassSerializer

    def perform_create(self, serializer):
        logger.info("Calling perform_create")
        logger.info(f"serializer.validated_data = {serializer.validated_data}")

        logger.info(
            "Popping rac_discount_code, discount_code, voucher_code, voucher_pin,\
                 concession_id, concession_cart_number and sold_via."
        )
        # Pop these values out so they don't mess with the model serializer
        rac_discount_code = serializer.validated_data.pop("rac_discount_code", None)
        discount_code = serializer.validated_data.pop("discount_code", None)
        voucher_code = serializer.validated_data.pop("voucher_code", None)
        voucher_pin = serializer.validated_data.pop("voucher_pin", None)
        concession_id = serializer.validated_data.pop("concession_id", None)
        concession_card_number = serializer.validated_data.pop(
            "concession_card_number", None
        )
        concession_card_expiry_month = serializer.validated_data.pop(
            "concession_card_expiry_month", None
        )
        concession_card_expiry_year = serializer.validated_data.pop(
            "concession_card_expiry_year", None
        )
        logger.debug(
            "concession_card_expiry_month = " + str(concession_card_expiry_month)
        )
        logger.debug(
            "concession_card_expiry_year = " + str(concession_card_expiry_year)
        )
        sold_via = serializer.validated_data.pop("sold_via", None)
        logger.info(
            "rac_discount_code, discount_code, voucher_code, voucher_pin, \
                concession_id, concession_cart_number and sold_via popped."
        )

        email_user_id = 0
        if is_retailer(self.request):
            logger.info(
                "This pass is being created by a retailer.",
            )
            # If the pass is being sold by a retailer, check if there is an existing email user
            # with the email address assigned to the pass
            email = serializer.validated_data["email"]
            if EmailUser.objects.filter(email=email).exists():
                logger.info(
                    f"User with email: {email} already exists in ledger.",
                )
                email_user = EmailUser.objects.get(email=email)
                email_user_id = email_user.id
                logger.info(
                    f"Calling serializer.save(user={email_user_id})",
                )
                park_pass = serializer.save(user=email_user_id)
                logger.info(
                    f"serializer.save(user={email_user_id}) called.",
                )
            else:
                logger.info(
                    "User with email does not exist in ledger.",
                )
                logger.info(
                    "Calling serializer.save()",
                )
                park_pass = serializer.save()
                logger.info(
                    "serializer.save() called.",
                )
        elif is_customer(self.request):
            logger.info(
                "This pass is being created by an external user.",
            )
            email_user_id = self.request.user.id
            logger.info(
                f"Calling serializer.save(user={email_user_id})",
            )
            park_pass = serializer.save(user=email_user_id)
            logger.info(
                f"serializer.save(user={email_user_id}) called.",
            )
        else:
            logger.warn(
                "This pass is being created by a user that is not a retailer or an external user.",
            )
            logger.info(
                "Calling serializer.save()",
            )
            park_pass = serializer.save()
            logger.info(
                "serializer.save() called.",
            )

        logger.info(f"Logging user action for: {park_pass}")
        UserAction.objects.log_action(
            object_id=park_pass.id,
            content_type=ContentType.objects.get_for_model(park_pass),
            who=email_user_id,
            what=settings.ACTION_CREATE.format(
                park_pass._meta.model.__name__, park_pass.id
            ),
        )
        logger.info("User action logged.")

        logger.info(f"Checking if sold_via: {sold_via} matches any retailer groups")
        if sold_via and RetailerGroup.objects.filter(id=sold_via).exists():
            retailer_group = RetailerGroup.objects.get(id=sold_via)
            logger.info(
                f"sold_via: {sold_via} matches retailer group: {retailer_group}",
            )
            park_pass.sold_via = retailer_group
            logger.info(
                f"sold_via: {sold_via} assigned to park pass: {park_pass}",
            )
        else:
            logger.info(
                f"sold_via: {sold_via} does not match any retailer groups",
            )
            park_pass.sold_via = RetailerGroup.get_dbca_retailer_group()
            logger.info(
                f"Default DBCA retailer group assigned to park pass: {park_pass}",
            )

        logger.info(f"Saving park pass: {park_pass}")
        park_pass.save()
        logger.info(f"Park pass: {park_pass} saved.")

        logger.info(
            f"Getting cart for user: {self.request.user.id} ({self.request.user})."
        )
        cart = Cart.get_or_create_cart(self.request)

        content_type = ContentType.objects.get_for_model(park_pass)
        oracle_code = CartUtils.get_oracle_code(
            self.request, content_type, park_pass.id
        )
        logger.info(f"Oracle code: {oracle_code} will be used for this park pass.")
        cart_item = CartItem(
            cart=cart,
            object_id=park_pass.id,
            content_type=content_type,
            oracle_code=oracle_code,
        )

        logger.info(f"Cart item: {cart_item} created in memory.")

        """ If the user deletes a cart item, any objects that can be attached to a cart item
        (concession usage, discount code usage and voucher transaction)
        are deleted in the cart item's delete method  """
        if rac_discount_code and check_rac_discount_hash(
            rac_discount_code, park_pass.email
        ):
            discount_percentage = Decimal(settings.RAC_DISCOUNT_PERCENTAGE)
            cart_item.rac_discount_usage = RACDiscountUsage.objects.create(
                park_pass=park_pass,
                discount_percentage=discount_percentage,
            )
        # Only check for concession if the user is not using an rac discount code
        elif concession_id and concession_card_number:
            if Concession.objects.filter(id=concession_id).exists():
                concession = Concession.objects.get(id=concession_id)
                logger.info(
                    f"This pass purchase includes a concession: {concession}.",
                )
                logger.info(
                    "Creating concession usage.",
                )
                last_month_before_expiry = timezone.datetime(
                    int(concession_card_expiry_year),
                    int(concession_card_expiry_month),
                    1,
                )
                # In most cases, the expiry date is the last day of the month
                # that is listed on the concession card so the real date of expiry is the
                # the first day of the next month.
                # I.e. if a card says expiry is 6/23 then it expires on the 1st day of 7/23 (at 12:00am midnight)
                concession_card_expiry = last_month_before_expiry + relativedelta(
                    months=1
                )

                concession_usage = ConcessionUsage.objects.create(
                    concession=concession,
                    park_pass=park_pass,
                    concession_card_number=concession_card_number,
                    date_expiry=concession_card_expiry,
                )
                logger.info(
                    "Concession usage: {concession_usage} created.",
                )
                cart_item.concession_usage = concession_usage
                logger.info(
                    "Concession usage assigned to cart item: {cart_item}.",
                )

        if discount_code:
            pass_type_id = park_pass.option.pricing_window.pass_type.id
            if DiscountCode.is_valid(discount_code, self.request.user.id, pass_type_id):
                discount_code = DiscountCode.objects.get(code=discount_code)
                logger.info(
                    f"This pass purchase includes a discount code: {discount_code}.",
                )
                logger.info(
                    "Creating discount code usage.",
                )
                discount_code_usage = DiscountCodeUsage.objects.create(
                    discount_code=discount_code, park_pass=park_pass
                )
                logger.info(
                    "Discount code usage: {discount_code_usage} created.",
                )
                cart_item.discount_code_usage = discount_code_usage
                logger.info(
                    "Discount code usage assigned to cart item: {cart_item}.",
                )

        if voucher_code:
            if Voucher.is_valid(voucher_code, voucher_pin):
                voucher = Voucher.objects.get(code=voucher_code, pin=voucher_pin)
                logger.info(
                    f"This pass purchase includes a voucher transaction for voucher: {voucher}.",
                )
                logger.info(
                    "Creating voucher transaction.",
                )
                voucher_transaction = VoucherTransaction.objects.create(
                    voucher=voucher,
                    park_pass=park_pass,
                    debit=voucher.balance_available_for_purchase(
                        park_pass.price_after_discount_code_applied
                    ),
                    credit=Decimal(0.00),
                )
                logger.info(
                    f"Voucher transaction: {voucher_transaction} created.",
                )
                cart_item.voucher_transaction = voucher_transaction
                logger.info(
                    f"Voucher transaction assigned to cart item: {cart_item}.",
                )

        # Save to apply any concession usages, discount code usages or voucher transactions to the cart item
        logger.info(f"Saving cart item: {cart_item}.")
        cart_item.save()
        logger.info(f"Cart item: {cart_item} saved.")

        if is_retailer(self.request) or is_customer(self.request):
            logger.info(
                "User is a retailer or external user so will increment cart item count.",
            )
            cart_item_count = CartUtils.increment_cart_item_count(self.request)
            logger.info(
                f"Incremented cart item count to {cart_item_count} -> {cart}",
            )

        if not cart.datetime_first_added_to:
            cart.datetime_first_added_to = timezone.now()
            logger.info(
                f"Assigned datetime_first_added_to {cart.datetime_first_added_to} for cart: {cart}.",
            )

        cart.datetime_last_added_to = timezone.now()
        logger.info(
            f"Assigned datetime_last_added_to {cart.datetime_last_added_to} for cart: {cart}."
        )

        logger.info(f"Saving cart: {cart}.")
        cart.save()
        logger.info(f"Cart: {cart} saved.")

    def perform_update(self, serializer):
        park_pass = serializer.save()
        UserAction.objects.log_action(
            object_id=park_pass.id,
            content_type=ContentType.objects.get_for_model(park_pass),
            who=self.request.user.id,
            what=settings.ACTION_UPDATE.format(
                park_pass._meta.model.__name__, park_pass.id
            ),
        )

    def has_object_permission(self, request, view, obj):
        # As per the drf docs, object level permissions are not applied when creating objects.
        if is_retailer(request):
            logger.info("Checking if retailer use have permission to access this pass.")
            retailer_group_ids_for_user = get_retailer_group_ids_for_user(
                request.user.id
            )
            if obj.sold_via in retailer_group_ids_for_user:
                return True
        if is_customer(request):
            if obj.user == request.user.id:
                return True
        return False

    @action(methods=["GET"], detail=True, url_path="retrieve-park-pass-pdf")
    def retrieve_park_pass_pdf(self, request, *args, **kwargs):
        park_pass = self.get_object()
        if park_pass.user == self.request.user.id:
            if park_pass.park_pass_pdf:
                return FileResponse(park_pass.park_pass_pdf)
        raise Http404

    @action(methods=["GET"], detail=True, url_path="retrieve-invoice")
    def retrieve_invoice(self, request, *args, **kwargs):
        park_pass = self.get_object()
        content_type = ContentType.objects.get_for_model(Pass)
        if OrderItem.objects.filter(
            object_id=park_pass.id, content_type=content_type
        ).exists():
            order_item = OrderItem.objects.get(
                object_id=park_pass.id, content_type=content_type
            )
            invoice_url = order_item.order.invoice_link
            logger.info(f"invoice_url: {invoice_url}")
            if invoice_url:
                response = requests.get(invoice_url)
                return FileResponse(response, content_type="application/pdf")

        raise Http404


class PassFilterBackend(DatatablesFilterBackend):
    """
    Custom Filters for Internal Pass Viewset
    """

    def filter_queryset(self, request, queryset, view):
        total_count = queryset.count()

        pass_type = request.GET.get("pass_type")
        status = request.GET.get("status")
        start_date_from = request.GET.get("start_date_from")
        start_date_to = request.GET.get("start_date_to")

        if pass_type:
            queryset = queryset.filter(option__pricing_window__pass_type__id=pass_type)

        if status:
            if Pass.CANCELLED == status:
                queryset = queryset.filter(cancellation__isnull=False)

            elif Pass.FUTURE == status:
                queryset = queryset.filter(date_start__gt=timezone.now().date())

            elif Pass.EXPIRED == status:
                queryset = queryset.filter(date_expiry__lte=timezone.now().date())

            elif Pass.CURRENT == status:
                queryset = queryset.exclude(processing_status=Pass.CANCELLED).filter(
                    date_start__lte=timezone.now().date(),
                    date_expiry__gte=timezone.now().date(),
                )

        if start_date_from:
            queryset = queryset.filter(date_start__gte=start_date_from)

        if start_date_to:
            queryset = queryset.filter(date_start__lte=start_date_to)

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        queryset = super().filter_queryset(request, queryset, view)
        setattr(view, "_datatables_total_count", total_count)

        return queryset


class RetailerPassViewSet(CustomDatatablesListMixin, UserActionViewSet):
    search_fields = [
        "pass_number",
        "first_name",
        "last_name",
        "vehicle_registration_1",
        "vehicle_registration_2",
    ]
    model = Pass
    pagination_class = DatatablesPageNumberPagination
    permission_classes = [IsRetailer]
    filter_backends = (PassFilterBackend,)
    renderer_classes = (JSONRenderer, BrowsableAPIRenderer, CustomDatatablesRenderer)
    http_method_names = ["get", "post", "head", "put", "patch"]

    def get_serializer_class(self):
        if "update" == self.action:
            return RetailerUpdatePassSerializer
        if "retrieve" == self.action:
            return InternalPassRetrieveSerializer
        return InternalPassSerializer

    def get_user_action_serializer_class(self):
        return UserActionSerializer

    def get_queryset(self):
        if RetailerGroupUser.objects.filter(
            emailuser__id=self.request.user.id
        ).exists():
            retailer_groups = RetailerGroupUser.objects.filter(
                emailuser__id=self.request.user.id
            ).values_list("retailer_group__id")
            return Pass.objects.exclude(in_cart=True).filter(
                sold_via__in=list(retailer_groups)
            )

        return Pass.objects.none()

    @action(methods=["GET"], detail=True, url_path="retrieve-park-pass-pdf")
    def retrieve_park_pass_pdf(self, request, *args, **kwargs):
        if RetailerGroupUser.objects.filter(
            emailuser__id=self.request.user.id
        ).exists():
            retailer_groups = RetailerGroupUser.objects.filter(
                emailuser__id=self.request.user.id
            ).values_list("retailer_group__id", flat=True)
            park_pass = self.get_object()
            if park_pass.sold_via.id in list(retailer_groups):
                if park_pass.park_pass_pdf:
                    return FileResponse(park_pass.park_pass_pdf)
        raise Http404


class InternalPassViewSet(CustomDatatablesListMixin, UserActionViewSet):
    search_fields = [
        "pass_number",
        "first_name",
        "last_name",
        "vehicle_registration_1",
        "vehicle_registration_2",
    ]
    model = Pass
    pagination_class = DatatablesPageNumberPagination
    queryset = Pass.objects.exclude(in_cart=True)
    permission_classes = [IsInternal]
    filter_backends = (PassFilterBackend,)
    renderer_classes = (JSONRenderer, BrowsableAPIRenderer, CustomDatatablesRenderer)

    def get_serializer_class(self):
        if "retrieve" == self.action:
            return InternalPassRetrieveSerializer
        return InternalPassSerializer

    def get_user_action_serializer_class(self):
        return UserActionSerializer

    @action(methods=["GET"], detail=True, url_path="retrieve-park-pass-pdf")
    def retrieve_park_pass_pdf(self, request, *args, **kwargs):
        park_pass = self.get_object()
        if park_pass.park_pass_pdf:
            return FileResponse(park_pass.park_pass_pdf)
        raise Http404

    @action(methods=["GET"], detail=True, url_path="payment-details")
    def payment_details(self, request, *args, **kwargs):
        park_pass = self.get_object()
        content_type = ContentType.objects.get_for_model(park_pass)
        if OrderItem.objects.filter(
            object_id=park_pass.id, content_type=content_type
        ).exists():
            order_item = OrderItem.objects.get(
                object_id=park_pass.id, content_type=content_type
            )
            invoice_reference = order_item.order.invoice_reference
            return redirect(
                settings.LEDGER_UI_URL
                + "/ledger/payments/oracle/payments?invoice_no="
                + invoice_reference
            )

        raise Http404

    @action(methods=["POST"], detail=True, url_path="pro-rata-refund")
    def pro_rata_refund(self, request, *args, **kwargs):
        park_pass = self.get_object()
        content_type = ContentType.objects.get_for_model(park_pass)
        if not OrderItem.objects.filter(
            content_type=content_type, object_id=park_pass.id
        ).exists():
            raise Http404
        orderitem = OrderItem.objects.get(
            content_type=content_type, object_id=park_pass.id
        )
        order = orderitem.order
        ledger_description = f"Pro-rata Refund of Pass: {park_pass.pass_number}"
        pro_rata_refund_amount = park_pass.pro_rata_refund_amount()
        if settings.DEBUG:
            pro_rata_refund_amount = int(pro_rata_refund_amount)
            ledger_description += " (Price rounded for dev env)"

        ledger_order_lines = [
            {
                "ledger_description": ledger_description,
                "quantity": 1,
                "price_incl_tax": str(-abs(pro_rata_refund_amount)),
                "oracle_code": CartUtils.get_oracle_code(
                    self.request, content_type, park_pass.id
                ),
                "line_status": settings.PARKPASSES_LEDGER_DEFAULT_LINE_STATUS,
            }
        ]
        booking_reference = str(uuid.uuid4())
        basket_parameters = CartUtils.get_basket_parameters(
            ledger_order_lines,
            booking_reference,
            is_no_payment=order.is_no_payment,
            booking_reference_link=order.uuid,
        )
        create_basket_session(request, request.user.id, basket_parameters)

        invoice_text = f"Park Passes Refund: {booking_reference}"
        return_url = request.build_absolute_uri(
            reverse(
                "internal-refund-success",
                kwargs={
                    "id": park_pass.id,
                    "uuid": booking_reference,
                },
            )
        )
        return_preload_url = request.build_absolute_uri(
            reverse(
                "ledger-api-refund-success-callback",
                kwargs={
                    "id": park_pass.id,
                    "uuid": booking_reference,
                },
            )
        )
        checkout_parameters = CartUtils.get_checkout_parameters(
            request, return_url, return_preload_url, order.user, invoice_text
        )
        logger.info("Checkout_parameters = " + str(checkout_parameters))

        create_checkout_session(request, checkout_parameters)

        return redirect(reverse("ledgergw-payment-details"))


class PassRefundSuccessView(APIView):
    def get(self, request, id, uuid, format=None):
        """We don't actually need to do any processing here but ledger needs a valid url for return_preload_url"""
        logger.info(f"RefundSuccessView get method called with id {id} and uuid {uuid}")


class PassAutoRenewSuccessView(APIView):
    def get(self, request, id, uuid, format=None):
        logger.info("Park passes Pass API PassAutoRenewSuccessView get method called.")

        invoice_reference = request.GET.get("invoice", "false")

        if id and uuid and invoice_reference:
            logger.info(
                f"New park pass id: {id}, Invoice reference: {invoice_reference} and uuid: {uuid}.",
            )
            if Pass.objects.filter(id=id).exists():
                new_park_pass = Pass.objects.get(id=id)
                logger.info(
                    f"Park pass with id={id} exists: {new_park_pass}.",
                )
                new_park_pass.in_cart = False
                new_park_pass.save()
                logger.info(
                    f"Park pass: {new_park_pass} removed from cart.",
                )

                if Order.objects.filter(uuid=uuid).exists():
                    order = Order.objects.get(uuid=uuid)
                    logger.info(
                        f"Order with uuid={uuid} exists: {order}",
                    )
                    order.invoice_reference = invoice_reference
                    logger.info(
                        f"Assigning invoice reference for: {order} to: {invoice_reference}",
                    )
                    order.save()
                    logger.info(
                        f"Invoice reference for: {order} assigned to: {invoice_reference}",
                    )
                else:
                    logger.error(f"Order with uuid: {uuid} does not exist.")

                logger.info(
                    f"Returning status.HTTP_204_NO_CONTENT. New Pass { new_park_pass }"
                    f" renewed successfully from { new_park_pass.park_pass_renewed_from }.",
                )
                # this end-point is called by an unmonitored get request in ledger so there is no point having a
                # a response body however we will return a status in case this is used on the ledger end in future
                return Response(status=status.HTTP_204_NO_CONTENT)

        logger.info(
            "Returning status.HTTP_400_BAD_REQUEST bad request as "
            "there was not a valid new park pass id, uuid and invoice_reference."
        )
        return Response(status=status.HTTP_400_BAD_REQUEST)


class CancelPass(APIView):
    permission_classes = [IsInternal]
    action = "cancel"

    def post(self, request, format=None):
        serializer = InternalPassCancellationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            park_pass = get_object_or_404(Pass, pk=serializer.data["park_pass"])
            user_action = settings.ACTION_CANCEL
            content_type = ContentType.objects.get_for_model(Pass)
            user_action = UserAction.objects.log_action(
                object_id=park_pass.id,
                content_type=content_type,
                who=request.user.id,
                what=user_action.format(Pass._meta.model.__name__, park_pass.id),
                why=serializer.data["cancellation_reason"],
            )
            extended_serializer = {
                "user_action": UserActionSerializer(user_action).data
            }
            extended_serializer.update(serializer.data)
            return Response(extended_serializer, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UploadPersonnelPasses(APIView):
    permission_classes = [IsInternal]
    action = "put"

    def put(self, request, format=None):
        logger.debug("files: " + str(request.FILES))
        personnel_data_file = request.FILES["personnelDataFile"]
        wrkbk = openpyxl.load_workbook(personnel_data_file, read_only=True)
        sh = wrkbk.active

        total_row_count = sh.max_row
        logger.info(f"total_row_count: {total_row_count}")

        if total_row_count <= 1:
            raise ValidationError(
                "The file you have uploaded doesn't have enough rows .xlsx file."
            )

        column_count = sh.max_column
        logger.info(f"column_count: {column_count}")
        if not column_count == 4:
            raise ValidationError(
                "The file .xlsx file needs to have 4 columns (First Name, Last Name, Email Address, Start Date)."
            )

        column_headings = []
        for cell in sh[1]:
            if cell.value:
                column_headings.append(cell.value)

        if "First Name" not in column_headings:
            raise ValidationError("'First Name' column must be present in .xlsx file.")

        if "Last Name" not in column_headings:
            raise ValidationError("'Last Name' column must be present in .xlsx file.")

        if "Email Address" not in column_headings:
            raise ValidationError(
                "'Email Address' column must be present in .xlsx file."
            )

        if "Start Date" not in column_headings:
            raise ValidationError("'Start Date' column must be present in .xlsx file.")

        logger.debug("Column Headings: " + str(column_headings))

        if not PassType.objects.filter(name=settings.PERSONNEL_PASS).exists():
            error = "Personnel Pass Type does not exist."
            logger.critical(error)
            raise ValidationError(error)

        pass_type = PassType.objects.get(name=settings.PERSONNEL_PASS)
        default_option = (
            PassTypePricingWindowOption.get_default_options_by_pass_type_id(
                pass_type.id
            ).first()
        )
        default_sold_via = RetailerGroup.get_dbca_retailer_group()

        park_passes_created = 0
        park_passes_duplicates = 0
        park_passes_errors = []

        data_row_count = total_row_count - 1

        # iterate through excel and display data
        for row in sh.iter_rows(min_row=2, max_col=4):
            logger.info(
                f"Attempting to create park pass for row: { row[0].value} {row[1].value} {row[2].value} {row[3].value}"
            )
            if Pass.objects.filter(
                first_name=row[0].value,
                last_name=row[1].value,
                email=row[2].value,
                date_start=row[3].value,
            ).exists():
                logger.warning(
                    f"Pass already exists for: \
                    { row[0].value} {row[1].value} {row[2].value} {row[3].value} - skipping row."
                )
                park_passes_duplicates += 1
                continue
            else:
                try:
                    aware_date = timezone.make_aware(row[3].value).date()
                    logger.debug("type(aware_date): " + str(type(aware_date)))
                    Pass.objects.create(
                        first_name=row[0].value,
                        last_name=row[1].value,
                        email=row[2].value,
                        date_start=aware_date,
                        option=default_option,
                        in_cart=False,
                        sold_via=default_sold_via,
                    )
                    park_passes_created += 1
                except Exception as e:
                    logger.error(e)
                    park_passes_errors.append(pickle.dumps(sys.exc_info()))

        results = {
            "data_row_count": data_row_count,
            "park_passes_created": park_passes_created,
            "park_passes_duplicates": park_passes_duplicates,
            "park_passes_errors": park_passes_errors,
        }

        logger.info("UploadPersonnelPasses results: " + str(results))

        return Response({"results": results}, status=status.HTTP_201_CREATED)


class RetailerApiAccessViewSet(UserActionViewSet):
    permission_classes = [HasRetailerGroupAPIKey]
    model = Pass
    pagination_class = DatatablesPageNumberPagination
    filter_backends = (PassFilterBackend,)
    http_method_names = ["get", "post", "put", "patch", "head", "options"]
    lookup_field = "rac_member_number"

    def get_serializer_class(self):
        if "retrieve" == self.action:
            return InternalPassRetrieveSerializer
        if "create" == self.action:
            return RetailerApiCreatePassSerializer
        return InternalPassSerializer

    def get_queryset(self):
        retailer_group = self.get_retailer_group(self.request)
        return (
            Pass.objects.exclude(in_cart=True)
            .filter(sold_via=retailer_group)
            .order_by("-datetime_created")
        )

    def perform_create(self, serializer):
        retailer_group = self.get_retailer_group(self.request)
        email = serializer.validated_data["email"]
        if EmailUser.objects.filter(email=email).exists():
            email_user = EmailUser.objects.get(email=email)
            serializer.save(user=email_user.id, sold_via=retailer_group)
        else:
            serializer.save(sold_via=retailer_group)

        return super().perform_create(serializer)

    @action(methods=["GET"], detail=True, url_path="get-discount-code-from-email")
    def get_discount_code_from_email(self, request, *args, **kwargs):
        email = kwargs["email"]
        regex = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
        if not re.fullmatch(regex, email):
            raise ValidationError({"email": "Please pass a valid email address."})
        return get_rac_discount_code(email)

    def get_retailer_group(self, request):
        """Retrieve a project based on the request API key."""
        if "HTTP_AUTHORIZATION" in request.META:
            key = request.META["HTTP_AUTHORIZATION"].split()[1]
            retailer_group_api_key = RetailerGroupAPIKey.objects.get_from_key(key)
            return retailer_group_api_key.retailer_group
        return RetailerGroup.get_rac_retailer_group()


class RacDiscountCodeView(APIView):
    permission_classes = [HasRetailerGroupAPIKey]

    def get(self, request, *args, **kwargs):
        email = kwargs["email"]
        regex = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
        if not re.fullmatch(regex, email):
            raise ValidationError({"email": "Please pass a valid email address."})
        return Response(get_rac_discount_code(email))

    def post(self, request, *args, **kwargs):
        regex = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
        emails = request.data.get("emails").split(",")
        codes = []
        for email in emails:
            if not re.fullmatch(regex, email):
                raise ValidationError(
                    {"email": f"Your list contains an invalid email address: {email}."}
                )
            codes.append({email: get_rac_discount_code(email)})
        return Response(codes)


class RacDiscountCodeCheckView(APIView):
    throttle_classes = [AnonRateThrottle]

    def get(self, request, *args, **kwargs):
        discount_hash = kwargs["discount_hash"]
        if 20 != len(discount_hash):
            raise ValidationError(
                {"discount_hash": "The discount hash must be 20 characters long."}
            )
        email = kwargs["email"]
        regex = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
        if not re.fullmatch(regex, email):
            raise ValidationError({"email": "Please pass a valid email address."})
        result = check_rac_discount_hash(discount_hash, email)
        if result:
            discount_percentage = Decimal(settings.RAC_DISCOUNT_PERCENTAGE)
            return Response(
                {
                    "is_rac_discount_code_valid": result,
                    "discount_percentage": discount_percentage,
                }
            )
        return Response({"is_rac_discount_code_valid": result})

    def get_retailer_group(self, request):
        """Retrieve a project based on the request API key."""
        if "HTTP_AUTHORIZATION" in request.META:
            key = request.META["HTTP_AUTHORIZATION"].split()[1]
            retailer_group_api_key = RetailerGroupAPIKey.objects.get_from_key(key)
            return retailer_group_api_key.retailer_group
        return RetailerGroup.get_rac_retailer_group()


class InternalDistrictPassTypeDurationOracleCodeViewSet(viewsets.ModelViewSet):
    model = DistrictPassTypeDurationOracleCode
    queryset = DistrictPassTypeDurationOracleCode.objects.all()
    permission_classes = [IsInternal]
    serializer_class = InternalDistrictPassTypeDurationOracleCodeSerializer
    renderer_classes = (JSONRenderer, BrowsableAPIRenderer)
    pagination_class = None

    def get_serializer_class(self):
        logger.info("action = %s", self.action)
        if "list_update" == self.action:
            return InternalDistrictPassTypeDurationOracleCodeListUpdateSerializer
        return InternalDistrictPassTypeDurationOracleCodeSerializer

    @action(methods=["PATCH"], detail=False, url_path="list-update")
    def list_update(self, request, *args, **kwargs):
        logger.info(
            "Calling list_update on InternalDistrictPassTypeDurationOracleCodeViewSet"
        )
        filter = request.data["filter"]
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.values_list("id", "oracle_code")
        if settings.PICA_ORACLE_CODE_LABEL == filter:
            queryset = queryset.filter(district__isnull=True)
        else:
            queryset = queryset.filter(district__name=filter)
        instances = list(queryset)
        logger.info("data = %s\n", str(request.data["data"]))

        data = request.data["data"]
        serializer = self.get_serializer(instances, data, many=True, partial=True)
        serializer.is_valid(raise_exception=True)
        logger.info("serializer.validated_data = %s\n", str(serializer.validated_data))
        self.perform_list_update(serializer)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_list_update(self, serializer):
        for instance_tuple, data in zip(serializer.instance, serializer.validated_data):
            id = instance_tuple[0]
            oracle_code = str(data["oracle_code"])
            instance = DistrictPassTypeDurationOracleCode.objects.get(id=id)
            if instance.oracle_code != oracle_code:
                logger.info("instance_tuple = %s", instance_tuple)
                logger.info("data = %s", data)
                logger.info("oracle_code = %s", oracle_code)
                instance.oracle_code = oracle_code
                instance.save()
